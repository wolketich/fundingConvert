from flask import Flask, request, send_file, render_template, jsonify
import pandas as pd
from datetime import datetime
import io
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
import os
import json

app = Flask(__name__)

# Function to parse allocation description and identify term/non-term times
def parse_allocation_description(description):
    hours, rate = description.split(" x ")
    hours = int(float(hours.split()[0]))  # Convert to float first and then to int
    rate = float(rate.replace("€", ""))
    return hours, rate

# Function to identify term and non-term times
def identify_term_non_term_times(hours_list):
    hours_list = sorted(set(hours_list))  # Remove duplicates and sort
    if len(hours_list) == 1:
        return str(hours_list[0])
    
    pairs = []
    used = set()
    for i in range(len(hours_list)):
        for j in range(i + 1, len(hours_list)):
            if abs(hours_list[j] - hours_list[i]) in [9, 12, 15] and hours_list[i] not in used and hours_list[j] not in used:
                pairs.append((hours_list[i], hours_list[j]))
                used.add(hours_list[i])
                used.add(hours_list[j])

    if pairs:
        pairs_str = [f"{pair[0]}/{pair[1]}" for pair in pairs]
        return ", ".join(pairs_str)

    remaining_hours = [str(hour) for hour in hours_list if hour not in used]
    return "-".join(remaining_hours)

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/upload', methods=['POST'])
def upload_file():
    funding_file = request.files['file-funding']
    chick_file = request.files['file-chick']
    children_file = request.files['file-children']
    
    df_funding = pd.read_excel(funding_file)
    df_chick = pd.read_excel(chick_file)
    df_children = pd.read_csv(children_file)

    # Filter CHICK data where all claims are confirmed by the parent
    df_chick = df_chick[(df_chick['All Claims Confirmed by Parent?'] == 'Yes')]

    # Transform the funding data
    transformed_data = []
    for index, row in df_funding.iterrows():
        child_name = row["Child"]
        allocation_date = datetime.strptime(row["Allocation Date"], "%d/%m/%Y")
        month = allocation_date.strftime("%b")
        # Skip rows where allocation dates are not between Aug 2024 and July 2025
        if allocation_date < datetime(2024, 8, 1) or allocation_date > datetime(2025, 7, 31):
            continue
        hours, rate = parse_allocation_description(row["Allocation Description"])
        allocation_value = row["Allocation Value"]
        
        transformed_data.append({
            "Child Name": child_name,
            "Month": month,
            "Hours": hours,
            "Rate": rate,
            "Allocation Value": allocation_value,
            "Allocation Date": allocation_date
        })

    # Create DataFrame for transformed data
    transformed_df = pd.DataFrame(transformed_data)

    # Group by child and month and summarize allocation values and hours
    grouped = transformed_df.groupby(["Child Name", "Month"]).agg({
        "Allocation Value": "sum",
        "Hours": list,
        "Rate": list,
        "Allocation Date": list
    }).reset_index()

    # Initialize the final output DataFrame
    months = ["Aug", "Sep", "Oct", "Nov", "Dec", "Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul"]
    output_columns = ["Name", "Date of Birth", "CHICK", "Claim Until", "Term/Non-Term/Changes", "Allocation Value"] + months + ["Child ID"]
    output_df = pd.DataFrame(columns=output_columns)

    unmatched_names = []
    matched_names = []
    children_dict = df_children.set_index('Full Name')['Child ID'].to_dict()

    # Populate the final output DataFrame
    for name, group in grouped.groupby("Child Name"):
        hours_info = identify_term_non_term_times(group["Hours"].explode().tolist())
        allocation_value = group["Allocation Value"].sum()
        chick_info = df_chick[df_chick['Child'] == name]
        if chick_info.empty:
            continue
        dob = chick_info['Date of Birth'].values[0]
        chick = chick_info['CHICK'].values[0]
        claim_until = chick_info['Claim Until'].values[0]
        row = {
            "Name": name,
            "Date of Birth": dob,
            "CHICK": chick,
            "Claim Until": claim_until,
            "Term/Non-Term/Changes": hours_info,
            "Allocation Value": f"€{allocation_value:.2f}"
        }
        for month in months:
            if month in group["Month"].values:
                row[month] = f"€{group[group['Month'] == month]['Allocation Value'].values[0]:.2f}"
            else:
                row[month] = ""
        if name in children_dict:
            row["Child ID"] = children_dict[name]
            matched_names.append(name)
        else:
            unmatched_names.append(name)
        output_df = pd.concat([output_df, pd.DataFrame([row])], ignore_index=True)

    # Exclude automatically matched names from the possible matches
    possible_matches = [match for match in df_children.to_dict(orient='records') if match['Full Name'] not in matched_names]

    # Save the DataFrame and unmatched names for manual matching
    output_df_file = "output_df.pkl"
    output_df.to_pickle(output_df_file)
    unmatched_file = "unmatched_names.json"
    unmatched_data = {"unmatched": sorted(unmatched_names), "possibleMatches": possible_matches}
    with open(unmatched_file, "w") as file:
        json.dump(unmatched_data, file)

    if unmatched_names:
        return jsonify(unmatched_data)

    # Convert DataFrame to Excel file with formatting
    output = io.BytesIO()
    workbook = Workbook()
    worksheet = workbook.active

    for r in dataframe_to_rows(output_df, index=False, header=True):
        worksheet.append(r)

    # Apply conditional formatting
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    amber_fill = PatternFill(start_color="FFBF00", end_color="FFBF00", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=4, max_col=4):
        for cell in row:
            if cell.value:
                claim_until_date = pd.to_datetime(cell.value, format="%d/%m/%Y", dayfirst=True)
                days_diff = (claim_until_date - datetime.now()).days
                if days_diff <= 7:
                    cell.fill = red_fill
                elif days_diff <= 14:
                    cell.fill = amber_fill
                elif days_diff <= 30:
                    cell.fill = yellow_fill

    workbook.save(output)
    output.seek(0)

    # Save the file temporarily to serve it later
    file_path = "funding_data_summary.xlsx"
    with open(file_path, "wb") as f_out:
        f_out.write(output.getvalue())

    return jsonify({"fileUrl": file_path})

@app.route('/finalize', methods=['POST'])
def finalize_matches():
    data = request.json
    matches = {item['name']: item['id'] for item in data['matches']}

    # Load the DataFrame and unmatched names
    output_df_file = "output_df.pkl"
    output_df = pd.read_pickle(output_df_file)
    unmatched_file = "unmatched_names.json"
    with open(unmatched_file, "r") as file:
        unmatched_data = json.load(file)

    # Update the DataFrame with the manual matches
    for name in unmatched_data["unmatched"]:
        if name in matches:
            output_df.loc[output_df['Name'] == name, 'Child ID'] = matches[name] if matches[name] else 0

    # Convert DataFrame to Excel file with formatting
    output = io.BytesIO()
    workbook = Workbook()
    worksheet = workbook.active

    for r in dataframe_to_rows(output_df, index=False, header=True):
        worksheet.append(r)

    # Apply conditional formatting
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    amber_fill = PatternFill(start_color="FFBF00", end_color="FFBF00", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=4, max_col=4):
        for cell in row:
            if cell.value:
                claim_until_date = pd.to_datetime(cell.value, format="%d/%m/%Y", dayfirst=True)
                days_diff = (claim_until_date - datetime.now()).days
                if days_diff <= 7:
                    cell.fill = red_fill
                elif days_diff <= 14:
                    cell.fill = amber_fill
                elif days_diff <= 30:
                    cell.fill = yellow_fill

    workbook.save(output)
    output.seek(0)

    # Save the file temporarily to serve it later
    file_path = "funding_data_summary.xlsx"
    with open(file_path, "wb") as f_out:
        f_out.write(output.getvalue())

    # Clean up temporary files
    os.remove(unmatched_file)
    os.remove(output_df_file)

    return jsonify({"fileUrl": file_path})

@app.route('/download/<filename>', methods=['GET'])
def download_file(filename):
    return send_file(filename, as_attachment=True)

if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5001)